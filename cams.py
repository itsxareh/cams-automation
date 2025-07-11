import time
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import numbers
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from dotenv import load_dotenv
import os 

load_dotenv()

USERNAME = os.getenv("USERLOGIN")
PASSWORD = os.getenv("PASSWORD")
CAMS_URL = os.getenv("CAMS_URL")

now = datetime.now()
timestamp = now.strftime("(%m-%d-%Y)_%I%M%S %p")
input_excel = "account_numbers.xlsx"
template_path = "(TEMPLATE).xlsx"


options = Options()
options.add_argument("--start-maximized")
options.add_argument("--ignore-certificate-errors")
options.add_argument("--ignore-ssl-errors")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
wait = WebDriverWait(driver, 10)

print("Opening login page...")
driver.get(CAMS_URL)

try:
    print("Waiting for login form...")
    username_field = wait.until(EC.presence_of_element_located((By.ID, "LoginID")))
    password_field = driver.find_element(By.ID, "txtPassword")
    login_button = driver.find_element(By.ID, "cmdLogin")
    
    print("Entering credentials...")
    username_field.clear()
    username_field.send_keys(USERNAME)
    password_field.clear()
    password_field.send_keys(PASSWORD)
    
    print("Logging in...")
    login_button.click()

    print("Waiting for login to process...")
    time.sleep(2) 
    
    current_url = driver.current_url
    page_source = driver.page_source
    
    if "login" in current_url.lower() or "error" in page_source.lower():
        print(f"Login may have failed - still on: {current_url}")
    
    print(f"Checking for new tabs... Current handles: {len(driver.window_handles)}")
    
    for i in range(10):
        if len(driver.window_handles) > 1:
            print(f"New tab detected after {i+1} seconds!")
            break
        time.sleep(.5)
    else:
        print("No new tab detected - login might have failed or uses different flow")
        print("Trying direct navigation to cams")
        driver.get("CAMS_URL")
        time.sleep(1)
        
        if "cams.aspx" in driver.current_url:
            print("Successfully accessed cams.aspx directly")
        else:
            print("Could not access main application page")
            with open("access_failed_dump.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            driver.quit()
            exit(1)
    
    if len(driver.window_handles) > 1:
        print("Switching to new tab...")
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(1.5)
        
        print(f"New tab URL: {driver.current_url}")
        print(f"New tab title: {driver.title}")
        
        if "cams.aspx" not in driver.current_url:
            print("New tab is not the expected cams.aspx page")
            print("Navigating to cams.aspx...")
            driver.get(CAMS_URL)
            time.sleep(1.25)
    
    print(f"✅ Ready to proceed - Current URL: {driver.current_url}")

except Exception as e:
    print(f"Login process error: {e}")
    driver.save_screenshot("login_error.png")
    driver.quit()
    raise e

try:
    time.sleep(1.25)
    
    frames = driver.find_elements(By.TAG_NAME, "frame")
    
    for i, frame in enumerate(frames):
        frame_name = frame.get_attribute("name")
        frame_src = frame.get_attribute("src")
        print(f"  Frame {i}: name='{frame_name}', src='{frame_src}'")
    
    print("Switching to SystemFrame...")
    try:
        driver.switch_to.frame("SystemFrame")
        print("Successfully switched to SystemFrame by name")
    except:
        try:
            system_frame = driver.find_element(By.NAME, "SystemFrame")
            driver.switch_to.frame(system_frame)
            print("Successfully switched to SystemFrame by element")
        except:
            driver.switch_to.frame(len(frames) - 1)
            print("Successfully switched to SystemFrame by index")
    
    time.sleep(1)
    
    print(f"Frame URL: {driver.current_url}")
    print(f"Frame title: {driver.title}")
    
    print("Looking for search elements in SystemFrame...")

except Exception as e:
    print(f"Failed to access frames: {e}")
    driver.save_screenshot("frame_error.png")
    with open("frame_error.html", "w", encoding="utf-8") as f:
        f.write(driver.page_source)
    driver.quit()
    raise e

print("📊 Loading account numbers from Excel...")
try:
    accounts_df = pd.read_excel(input_excel, dtype={"Account No": str})
    account_dict = accounts_df.set_index("Account No").to_dict("index")
    account_numbers = list(account_dict.keys())
    print(f"📋 Loaded {len(account_numbers)} account numbers")
except Exception as e:
    print(f"Error loading Excel file: {e}")
    driver.quit()
    raise e

results = []

print("Starting account processing...")

print("Ensuring we're in the SystemFrame...")
try:
    driver.switch_to.default_content()
    driver.switch_to.frame("SystemFrame")
    print("Switched to SystemFrame for element checking")
except Exception as e:
    print(f"Frame switching warning: {e}")

print("Checking if search elements are available in SystemFrame...")

def check_element_exists(by, identifier, element_name):
    """Check if an element exists and return True/False with logging"""
    try:
        element = driver.find_element(by, identifier)
        print(f"Found {element_name}: {identifier}")
        return True
    except:
        print(f"Missing {element_name}: {identifier}")
        return False

search_elements_found = True
if not check_element_exists(By.ID, "_ctl0_ContentPlaceHolder1_txtSearch", "Search Input"):
    search_elements_found = False
if not check_element_exists(By.ID, "_ctl0_ContentPlaceHolder1_btnSearch", "Search Button"):
    search_elements_found = False

if not search_elements_found:
    print("Critical search elements not found. Saving page state for debugging...")
    driver.save_screenshot("missing_search_elements.png")
    with open("search_elements_missing.html", "w", encoding="utf-8") as f:
        f.write(driver.page_source)
    
    print("Looking for alternative search elements...")
    
    all_inputs = driver.find_elements(By.TAG_NAME, "input")
    print(f"Found {len(all_inputs)} input elements:")
    for i, inp in enumerate(all_inputs[:10]):  
        try:
            inp_id = inp.get_attribute("id")
            inp_name = inp.get_attribute("name")
            inp_type = inp.get_attribute("type")
            inp_placeholder = inp.get_attribute("placeholder")
            if inp_id or inp_name:
                print(f"  [{i}] ID: {inp_id}, Name: {inp_name}, Type: {inp_type}, Placeholder: {inp_placeholder}")
        except:
            pass
    
    all_buttons = driver.find_elements(By.TAG_NAME, "button")
    all_inputs_button = driver.find_elements(By.XPATH, "//input[@type='button' or @type='submit']")
    all_buttons.extend(all_inputs_button)
    print(f"Found {len(all_buttons)} button elements:")
    for i, btn in enumerate(all_buttons[:10]): 
        try:
            btn_id = btn.get_attribute("id")
            btn_name = btn.get_attribute("name")
            btn_value = btn.get_attribute("value")
            btn_text = btn.text
            if btn_id or btn_name or btn_value or btn_text:
                print(f"  [{i}] ID: {btn_id}, Name: {btn_name}, Value: {btn_value}, Text: {btn_text}")
        except:
            pass

if not search_elements_found:
    print("Cannot proceed without search elements. Please check the page structure.")
    driver.quit()
    exit(1)

for i, acct in enumerate(account_numbers, 1):
    print(f"[{i}/{len(account_numbers)}] Searching for account: {acct}")
    
    try:
        manual = account_dict.get(acct, {})
        
        print(f"Navigating to main search page for account {acct}")
        driver.get(CAMS_URL)
        time.sleep(1)  
        
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame("SystemFrame")
            print(f"Switched to SystemFrame for account {acct}")
        except Exception as frame_error:
            print(f"Frame switching error for {acct}: {frame_error}")
            continue
        
        try:
            search_input = wait.until(EC.presence_of_element_located((By.ID, "_ctl0_ContentPlaceHolder1_txtSearch")))
            print(f"Search input found for account {acct}")
        except:
            print(f"Search input not found for account {acct}. Skipping...")
            driver.save_screenshot(f"search_input_missing_{acct.replace('/', '_')}.png")
            continue
        
        search_input.clear()
        search_input.send_keys(acct)
        print(f"Entered account number: {acct}")
        
        try:
            search_button = driver.find_element(By.ID, "_ctl0_ContentPlaceHolder1_btnSearch")
            search_button.click()
            print(f"Search button clicked for {acct}")
        except:
            print(f"Search button not found for account {acct}. Skipping...")
            driver.save_screenshot(f"search_button_missing_{acct.replace('/', '_')}.png")
            continue
        
        time.sleep(.25)
        
        results_area_selectors = [
            f"//td[contains(text(), '{acct}')]",
            f"//tr[contains(., '{acct}')]",
            f"//*[contains(text(), '{acct}')]"
        ]
        
        account_link = None
        for selector in results_area_selectors:
            try:
                short_wait = WebDriverWait(driver, .25)
                account_link = short_wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                print(f"Found account result with selector: {selector}")
                break
            except:
                continue
        
        if account_link:
            account_link.click()
            print(f"Found and selected account: {acct}")
            time.sleep(.25)
            
            try:
                popup_ok_selectors = [
                    (By.XPATH, "//button[text()='OK']"),
                    (By.XPATH, "//button[contains(text(), 'OK')]"),
                    (By.XPATH, "//input[@value='OK']"),
                    (By.ID, "OK"),
                    (By.CLASS_NAME, "ui-button"),
                    (By.XPATH, "//div[contains(@class, 'ui-dialog')]//button"),
                    (By.XPATH, "//button[contains(@class, 'ylin') and contains(text(), 'OK')]")
                ]
                
                popup_dismissed = False
                for by, selector in popup_ok_selectors:
                    try:
                        popup_wait = WebDriverWait(driver, .75)
                        ok_button = popup_wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                        ok_button.click()
                        print(f"Dismissed popup for account {acct}")
                        popup_dismissed = True
                        time.sleep(.75)  
                        break
                    except:
                        continue
                
                if not popup_dismissed:
                    print(f"No popup found for account {acct} (this is normal)")
                    
            except Exception as popup_error:
                print(f"Error handling popup for {acct}: {popup_error}")
            
        else:
            print(f"Account {acct} not found in search results - skipping...")
            continue

        def safe_text(by, ident):
            try:
                el = driver.find_element(by, ident)
                value = el.get_attribute("value") or el.text
                return value.strip() if value else ""
            except Exception as e:
                print(f"Could not find element {ident}: {str(e)}")
                return ""

        print(f"Extracting account details for {acct}...")
        
        rem_stat_full = safe_text(By.ID, "txtREM_STAT")
        rem_stat_number = rem_stat_full.split('/')[0].strip() if rem_stat_full else ""
        
        endo_date = manual.get("ENDO DATE", "")
        if endo_date:
            try:
                endo_date = pd.to_datetime(endo_date).strftime("%m/%d/%Y")
            except:
                print(f"Invalid ENDO DATE format for {acct}: {endo_date}")
        
        new_pullout_date = manual.get("NEW_PULLOUT_DATE", "")
        if new_pullout_date:
            try:
                new_pullout_date = pd.to_datetime(new_pullout_date).strftime("%m/%d/%Y")
            except:
                print(f"Invalid NEW_PULLOUT_DATE format for {acct}: {new_pullout_date}")
        
        data = {
            "AGENT": manual.get("AGENT", ""),
            "ENDO DATE": endo_date,
            "PLACEMENT": manual.get("PLACEMENT", ""),
            "ACCOUNT NO": acct,
            "ACCOUNT NAME": safe_text(By.ID, "lblSHORTNAME"),
            "PRIMARY ADDRESS": safe_text(By.ID, "lblADDRESS"),
            "SECONDARY ADDRESS": safe_text(By.ID, "lblADDRESS2"),
            "OB": safe_text(By.ID, "lblOUT_BAL"),
            "REM STAT": rem_stat_number,
            "DPD": safe_text(By.ID, "lblAGESRC"),
            "MOBILE #": safe_text(By.ID, "txtMobile"),
            "NEW_PULLOUT_DATE": new_pullout_date,
            "Status": "Success"
        }
        
        if not any(data[key] for key in data if key not in ["ACCOUNT NO", "Status", "AGENT", "ENDO DATE", "PLACEMENT", "NEW_PULLOUT_DATE"]):
            print(f"No basic account information found for {acct}. Skipping...")
            driver.save_screenshot(f"no_account_info_{acct.replace('/', '_')}.png")
            continue

        try:
            account_details_selectors = [
                (By.PARTIAL_LINK_TEXT, "Account Details"),
                (By.LINK_TEXT, "Account Details"),
                (By.XPATH, "//a[contains(text(), 'Account Details')]"),
                (By.CLASS_NAME, "dropdown-toggle"),
                (By.XPATH, "//a[@class='dropdown-toggle' and contains(text(), 'Account Details')]")
            ]
            
            account_details_dropdown = None
            for by, selector in account_details_selectors:
                try:
                    account_details_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                    print(f"Found Account Details dropdown with: {selector}")
                    break
                except:
                    continue
            
            if account_details_dropdown:
                account_details_dropdown.click()
                print(f"Opened Account Details dropdown for {acct}")
                time.sleep(.75)
                
                collateral_selectors = [
                    (By.PARTIAL_LINK_TEXT, "Collateral"),
                    (By.LINK_TEXT, "Collateral"),
                    (By.XPATH, "//a[contains(text(), 'Collateral')]"),
                    (By.XPATH, "//a[contains(@href, 'accountcollateral.aspx')]"),
                    (By.XPATH, "//*[contains(text(), 'Collateral')]")
                ]
                
                collateral_tab = None
                for by, selector in collateral_selectors:
                    try:
                        collateral_tab = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                        print(f"Found collateral link with: {selector}")
                        break
                    except:
                        continue
                
                if collateral_tab:
                    collateral_tab.click()
                    print(f"Accessing collateral information for {acct}")
                    time.sleep(1)
                    
                    try:
                        vehicle_detail_selectors = [
                            "dg__ctl2_DetailLink",
                            "DetailLink",
                            "//a[contains(@id, 'DetailLink')]",
                            "//a[contains(text(), 'Detail')]",
                            "//a[contains(@id, 'dg_') and contains(@id, 'DetailLink')]"
                        ]
                        
                        vehicle_detail_link = None
                        for selector in vehicle_detail_selectors:
                            try:
                                if selector.startswith("//"):
                                    vehicle_detail_link = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                                else:
                                    vehicle_detail_link = wait.until(EC.element_to_be_clickable((By.ID, selector)))
                                print(f"Found vehicle detail link: {selector}")
                                break
                            except:
                                continue
                        
                        if vehicle_detail_link:
                            vehicle_detail_link.click()
                            time.sleep(1)
                            
                            vehicle_data = {
                                "COLOR": safe_text(By.ID, "txtColor_AUTO"),
                                "PLATE #": safe_text(By.ID, "txtPlateNum_AUTO"),
                                "SERIAL #": safe_text(By.ID, "txtSerialNo_AUTO"),
                                "ENGINE #": safe_text(By.ID, "txtEngineNo_AUTO"),
                                "UNIT DESCRIPTION": safe_text(By.ID, "txtUnitDesc_AUTO"),
                            }
                            data.update(vehicle_data)
                            
                            if any(vehicle_data.values()):
                                print(f"Vehicle details extracted for {acct}")
                            else:
                                print(f"No vehicle details found for {acct}")
                        else:
                            print(f"Vehicle detail link not found for {acct}")
                            
                    except Exception as e:
                        print(f"Could not access vehicle details for {acct}: {e}")
                        
                else:
                    print(f"Collateral link not found in dropdown for {acct}")
            else:
                print(f"Account Details dropdown not found for {acct}")
                
        except Exception as e:
            print(f"Could not access Account Details dropdown for {acct}: {e}")

        # Only append successful account data
        data["Status"] = "Success"
        results.append(data)
        print(f"Successfully processed account: {acct}")

    except Exception as e:
        print(f"Error processing account {acct}: {str(e)}")
        continue
    
    if i < len(account_numbers):  
        print(f"Waiting before processing next account...")
        time.sleep(.5)

print("Saving results to Excel using template...")
try:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    headers = {}
    for cell in ws[1]:
        if cell.value:
            header_name = str(cell.value).strip().upper()
            headers[header_name] = cell.column
            print(f"Found header: '{header_name}' in column {cell.column}")
    
    print(f"Template headers found: {list(headers.keys())}")
    
    row_idx = 2 
    for entry in results:
        if entry.get("Status") != "Success":
            continue
            
        print(f"Writing data for row {row_idx}: Account {entry.get('ACCOUNT NO', 'Unknown')}")
        
        for data_key, data_value in entry.items():
            if data_key == "Status":
                continue
                
            data_key_upper = str(data_key).strip().upper()
            
            if data_key_upper in headers:
                col_num = headers[data_key_upper]
                cell = ws.cell(row=row_idx, column=col_num)
                cell.value = str(data_value) if data_value is not None else ""
                if data_key_upper in ["ENDO DATE", "NEW_PULLOUT_DATE"]:
                    cell.number_format = '@'
                print(f"Set {data_key_upper}: '{data_value}' in column {col_num}")
            else:
                print(f"No matching header found for: '{data_key_upper}'")
        
        row_idx += 1

    output_excel = f"psb_auto-new-{timestamp}.xlsx"

    wb.save(output_excel)
    print(f"Results saved to: {output_excel}")
    print(f"Total accounts processed: {len(results)}")
    
    successful = sum(1 for r in results if r.get("Status") == "Success")
    not_found = sum(1 for r in results if r.get("Status") == "Not Found")
    errors = len(results) - successful - not_found
    
    print(f"Summary:")
    print(f"  - Successful: {successful}")
    print(f"  - Not Found: {not_found}")
    print(f"  - Errors: {errors}")
    
except Exception as e:
    print(f"Error saving results: {e}")
    try:
        results_df = pd.DataFrame([r for r in results if r.get("Status") == "Success"])
        results_df = results_df.drop(columns=["Status"], errors='ignore')
        if "ENDO DATE" in results_df.columns:
            results_df["ENDO DATE"] = pd.to_datetime(results_df["ENDO DATE"], errors='coerce').dt.strftime("%m/%d/%Y")
        if "NEW_PULLOUT_DATE" in results_df.columns:
            results_df["NEW_PULLOUT_DATE"] = pd.to_datetime(results_df["NEW_PULLOUT_DATE"], errors='coerce').dt.strftime("%m/%d/%Y")
        fallback_filename = f"psb_auto-fallback-{timestamp}.xlsx"
        results_df.to_excel(fallback_filename, index=False)
        print(f"Fallback file saved: {fallback_filename}")
    except Exception as fallback_error:
        print(f"Fallback save also failed: {fallback_error}")

print("\n Process completed!")
#input("Press ENTER to close the browser...")
driver.quit()
